#!/usr/bin/env python3
"""Extract PlanU course catalog JSON from a PNU timetable workbook."""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_ALIASES = {
    "course_id": ["교과목번호", "과목번호"],
    "course_name": ["교과목명(미확정구분)", "교과목명", "과목명"],
    "section": ["분반"],
    "credits": ["학점"],
    "instructor": ["교수명", "담당교수", "담당"],
    "time_room": ["시간/강의실", "시간", "강의시간"],
}

TIME_ROOM_RE = re.compile(
    r"(?P<day>[월화수목금토일])\s+"
    r"(?P<start>\d{1,2}:\d{2})"
    r"(?:(?:\((?P<minutes>\d+)\))|(?:-(?P<end>\d{1,2}:\d{2})))?\s+"
    r"(?P<room>[0-9A-Za-z가-힣\-]+)"
)


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_column(header: list[str], aliases: list[str]) -> int | None:
    for alias in aliases:
        if alias in header:
            return header.index(alias)
    return None


def add_minutes(time_text: str, minutes: int) -> str:
    hour_text, minute_text = time_text.split(":", 1)
    total = int(hour_text) * 60 + int(minute_text) + minutes
    return f"{total // 60:02d}:{total % 60:02d}"


def normalize_section(value: Any) -> str:
    text = cell_text(value)
    if not text:
        return "미확인"
    if text.isdigit():
        return text.zfill(3)
    return text


def split_time_room(value: Any) -> list[dict[str, str | int | None]]:
    text = unescape(cell_text(value))
    if not text:
        return [
            {
                "day": "미확인",
                "period": "미확인",
                "start_time": None,
                "end_time": None,
                "building": "미확인",
                "room": "미확인",
                "campus_area": "미확인",
            }
        ]

    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    chunks = [part.strip(" ,") for part in re.split(r"[\n,]+", text) if part.strip(" ,")]
    blocks: list[dict[str, str | int | None]] = []

    for chunk in chunks:
        match = TIME_ROOM_RE.search(chunk)
        if not match:
            blocks.append(
                {
                    "day": "미확인",
                    "period": "미확인",
                    "start_time": None,
                    "end_time": None,
                    "building": "미확인",
                    "room": chunk or "미확인",
                    "campus_area": "미확인",
                }
            )
            continue

        end_match = match.group("end")
        minutes = int(match.group("minutes") or 75)
        room = match.group("room")
        building = room.split("-", 1)[0] if "-" in room else room
        campus_area = building[0] if building and building[0].isdigit() else "미확인"
        start_time = match.group("start")
        end_time = end_match or add_minutes(start_time, minutes)
        blocks.append(
            {
                "day": match.group("day"),
                "period": f"{start_time}-{end_match}" if end_match else f"{start_time}({minutes})",
                "start_time": start_time,
                "end_time": end_time,
                "building": building or "미확인",
                "room": room or "미확인",
                "campus_area": campus_area,
            }
        )

    return blocks


def detect_header(ws) -> tuple[int, list[str]]:
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        header = [cell_text(value) for value in row]
        if "분반" in header and any(name in header for name in HEADER_ALIASES["course_name"]):
            return row_index, header
    raise ValueError("헤더 행을 찾지 못했습니다. 열 이름과 첫 3~5줄 샘플이 필요합니다.")


def extract(path: Path, sheet_name: str | None = None) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    header_row, header = detect_header(ws)

    columns = {
        key: find_column(header, aliases)
        for key, aliases in HEADER_ALIASES.items()
    }
    missing = [key for key, index in columns.items() if index is None]
    if missing:
        raise ValueError(f"필수 열을 찾지 못했습니다: {', '.join(missing)}")

    courses: list[dict[str, Any]] = []
    warnings: list[str] = []

    for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        course_name = cell_text(row[columns["course_name"]])  # type: ignore[index]
        if not course_name:
            continue

        course_id = cell_text(row[columns["course_id"]]) or "미확인"  # type: ignore[index]
        section = normalize_section(row[columns["section"]])  # type: ignore[index]
        instructor = cell_text(row[columns["instructor"]]) or "미확인"  # type: ignore[index]
        credits_raw = row[columns["credits"]]  # type: ignore[index]
        try:
            credits = int(credits_raw) if credits_raw not in (None, "") else None
        except (TypeError, ValueError):
            credits = None
            warnings.append(f"{row_index}행 학점 값을 숫자로 변환하지 못했습니다.")

        for block in split_time_room(row[columns["time_room"]]):  # type: ignore[index]
            courses.append(
                {
                    "course_id": course_id,
                    "course_name": course_name,
                    "section": section,
                    "credits": credits,
                    "day": block["day"],
                    "period": block["period"],
                    "start_time": block["start_time"],
                    "end_time": block["end_time"],
                    "building": block["building"],
                    "room": block["room"],
                    "instructor": instructor,
                    "campus_area": block["campus_area"],
                    "source_row": row_index,
                    "evidence": f"{path.name}:{ws.title}:{row_index}",
                }
            )

    return {
        "source_file": path.name,
        "extracted_at": datetime.now().replace(microsecond=0).isoformat(),
        "courses": courses,
        "warnings": warnings,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract PNU course catalog JSON.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default=None)
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"파일을 찾지 못했습니다: {args.workbook}", file=sys.stderr)
        return 2

    try:
        result = extract(args.workbook, args.sheet)
    except Exception as exc:
        print(f"추출 실패: {exc}", file=sys.stderr)
        return 1

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
